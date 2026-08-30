"""gera planilha de apresentacao JULHO x AGOSTO com dados extraidos direto da
planilha de KM (via processar_planilha + preparar). Cria 4 abas: Resumo,
Combustivel, Veiculos, Meses e Textos. Os textos e numeros sao derivados dos
dados (nada de valores fixos) — exceto os comentarios qualitativos.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from frota_utils import processar_planilha
from gerar_html import preparar

DIR = Path(__file__).resolve().parent
ARQUIVO_KM = DIR / "PA - CONTROLE DE KM (version 1).xlsx"
SAIDA = DIR / "Apresentacao_Julho_Agosto.xlsx"

AZUL = "2563EB"
AMARELO = "F59E0B"
VERDE = "10B981"
VERMELHO = "EF4444"
CINZA = "64748B"
BANCO = "111C30"

THIN = Side(style="thin", color="D1D9E4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TIT = Font(bold=True, size=16, color="FFFFFF")
F_SEG = Font(bold=True, size=13, color="22303F")
F_HED = Font(bold=True, size=11, color="FFFFFF")
F_BLD = Font(bold=True, color="22303F")
F_POS = Font(color=VERDE, bold=True)
F_NEG = Font(color=VERMELHO, bold=True)
FILL_TIT = PatternFill("solid", fgColor=AZUL)
FILL_HED = PatternFill("solid", fgColor="14507A")
FILL_ZEB = PatternFill("solid", fgColor="F1F5FB")
AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def extrair_dados() -> dict:
    """Monta o mesmo dicionario que alimenta o dashboard HTML, mas direto do
    workbook (sem depender de parse do arquivo gerado)."""
    df = processar_planilha(ARQUIVO_KM)
    if df.empty:
        raise SystemExit("Nenhuma aba mensal válida na planilha de KM.")
    return preparar(df)


def br(v, dec=1) -> str:
    """Formata número pt-BR com milhar por ponto e decimal por vírgula."""
    if v is None:
        return "—"
    s = f"{v:,.{dec}f}"
    return s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def d(v, dec=1) -> str:
    """Variação com sinal e valor absoluto ('−6.517' / '+6,0')."""
    return ("+" if v >= 0 else "−") + br(abs(v), dec)


def pct(v, dec=1) -> str:
    """Percentual com sinal ('−33,9%')."""
    return d(v, dec) + "%"


def numerico(v) -> str:
    if v is None:
        return "—"
    return br(v, 2)


def f_km(v) -> str:
    return br(v, 0) + " km"


def estilo_cabecalho(ws, linha: int, n_col: int):
    for c in range(1, n_col + 1):
        cel = ws.cell(row=linha, column=c)
        cel.font = F_HED
        cel.fill = FILL_HED
        cel.alignment = AL_C
        cel.border = BORDER


def zebrar(ws, r_inicio: int, r_fim: int, n_col: int):
    for r in range(r_inicio, r_fim + 1):
        if (r - r_inicio) % 2 == 1:
            for c in range(1, n_col + 1):
                ws.cell(row=r, column=c).fill = FILL_ZEB


def linha_total(ws, r: int, n_col: int, label: str = "TOTAL"):
    for c in range(1, n_col + 1):
        cel = ws.cell(row=r, column=c)
        cel.font = F_BLD
        cel.border = BORDER
    ws.cell(row=r, column=1).value = label


def pct_num(p) -> float | None:
    """Converte % do JSON (ex: -33.87) em valor fracionario p/ formatar como %."""
    return None if p is None else p / 100.0


def aplicar_delta(cel, delta: float | None, bom_negativo: bool = True):
    """Pinta a celula de delta conforme semantica: verde quando (nao) foi bom."""
    if delta is None:
        cel.value = "—"
        return
    up = delta > 0
    bom = (not up) if bom_negativo else up
    cel.value = f"{'' if up else '−'}{numerico(abs(delta))}"
    cel.font = F_POS if bom else F_NEG


def km_of(r):
    return r[3] or 0


def co2_of(r):
    return r[5] or 0


def main() -> None:
    dados = extrair_dados()
    consumo = {c["mes"]: c for c in dados["consumoTend"]}
    jul = consumo["JULHO"]
    ago = consumo["AGOSTO"]
    pj = dados["por_mes"]["JULHO"]
    pa = dados["por_mes"]["AGOSTO"]
    a_by = {r[0]: r for r in dados["consumo"]["AGOSTO"]}
    j_by = {r[0]: r for r in dados["consumo"]["JULHO"]}

    km_j, km_a = jul["km"], ago["km"]
    veic_j, veic_a = jul["veic"], ago["veic"]
    lit_j = jul["litrosD"] + jul["litrosG"]
    lit_a = ago["litrosD"] + ago["litrosG"]
    co2_j, co2_a = jul["total"], ago["total"]
    med_j = km_j / veic_j if veic_j else 0
    med_a = km_a / veic_a if veic_a else 0

    km01_j = km_of(j_by.get("TN 01"))
    km01_a = km_of(a_by.get("TN 01"))
    top_ago = max(a_by.values(), key=km_of) if a_by else None
    ts = top_ago[0] if top_ago else ""
    ts_kj = km_of(j_by.get(ts))
    ts_ka = km_of(top_ago)

    if "TN 76" not in a_by and "TN 76" in j_by and km_of(j_by["TN 76"]) == 0:
        tn76_txt = " TN 76 saiu da operação (0 km em JULHO e sem registro em AGOSTO)."
    elif "TN 76" not in a_by:
        tn76_txt = " TN 76 não aparece em AGOSTO."
    else:
        tn76_txt = ""

    delta01 = km01_a - km01_j
    if delta01 > 0:
        aumento_txt = (f"O que aumentou: apenas a gasolina do veículo leve "
                       f"(TN 01), que rodou {d(delta01, 0)} km no mês.")
    else:
        aumento_txt = ("O único movimento relevante em combustível foi o veículo "
                       "leve (TN 01), que roda a Gasolina.")

    p_diesel_j = dc_j / co2_j * 100 if (dc_j := jul["diesel"]) and co2_j else 0
    p_diesel_a = dc_a / co2_a * 100 if (dc_a := ago["diesel"]) and co2_a else 0

    top5 = sorted(a_by.values(), key=co2_of, reverse=True)[:5]
    top_txt = "; ".join(
        (f"{r[0]} com {br(co2_of(r), 1)} kg"
         + (f" ({pct((co2_of(r) - co2_of(j_by[r[0]])) / (co2_of(j_by[r[0]]) or 1) * 100)})"
            if r[0] in j_by and co2_of(j_by[r[0]]) else ""))
        for r in top5)
    top0 = top5[0] if top5 else None
    pct_top = (co2_of(top0) / co2_a * 100) if top0 and co2_a else 0
    diff_top = (co2_of(j_by[top0[0]]) - co2_of(top0)) \
        if top0 and top0[0] in j_by else None

    km_d = sum(km_of(r) for r in a_by.values() if r[2] == "Diesel")
    pct_km_diesel = km_d / km_a * 100 if km_a else 0
    reds = [(t, km_of(j_by[t]) - km_of(r)) for t, r in a_by.items()
            if t in j_by and km_of(j_by[t]) > km_of(r)]
    reds.sort(key=lambda x: -x[1])
    ex_reducoes = "; ".join(f"{t} {d(-dk, 0)} km" for t, dk in reds[:3]) \
        or "todas as caminhonetes rodaram menos"

    wb = Workbook()

    total_v = dict(zip(dados["meses"], dados["tendencia"]))  # pragma: no cover

    # =====================================================================
    # ABA 1 — RESUMO
    # =====================================================================
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:H1")
    ws["A1"] = "RESUMO — JULHO × AGOSTO (dados do Dashboard de Frota Carajás)"
    ws["A1"].font = F_TIT
    ws["A1"].fill = FILL_TIT
    ws["A1"].alignment = AL_C
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Métrica"
    ws["B3"] = "JULHO"
    ws["C3"] = "AGOSTO"
    ws["D3"] = "Δ (absoluto)"
    ws["E3"] = "Δ (%)"
    estilo_cabecalho(ws, 3, 5)

    met = [
        ("KM total", f_km(km_j), f_km(km_a), km_a - km_j, (km_a - km_j) / km_j),
        ("Veículos no mês", veic_j, veic_a, veic_a - veic_j, (veic_a - veic_j) / veic_j),
        ("Litros totais (L)", numerico(lit_j), numerico(lit_a),
         lit_a - lit_j, (lit_a - lit_j) / lit_j),
        ("CO₂ total (kgCO₂e)", numerico(co2_j), numerico(co2_a),
         co2_a - co2_j, (co2_a - co2_j) / co2_j),
        ("Média por veículo (km)", f_km(round(med_j)), f_km(round(med_a)),
         med_a - med_j, (med_a - med_j) / med_j),
    ]
    r = 4
    for nome, vj, va, delta, frac in met:
        ws.cell(row=r, column=1, value=nome).font = F_BLD
        ws.cell(row=r, column=2, value=vj)
        ws.cell(row=r, column=3, value=va)
        if delta is not None:
            aplicar_delta(ws.cell(row=r, column=4), delta)
            ws.cell(row=r, column=5).value = frac
            ws.cell(row=r, column=5).number_format = "+0.0%;-0.0%"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            if c in (2, 3):
                ws.cell(row=r, column=c).alignment = AL_C
        r += 1

    ws["A9"] = "Frota operacional"
    ws["B9"] = f"{pj['ativos']} de {pj['veiculos']} ({pj['pctOp']:.0f}%)"
    ws["C9"] = f"{pa['ativos']} de {pa['veiculos']} ({pa['pctOp']:.0f}%)"
    ws["A10"] = "Em manutenção"
    ws["B10"] = pj["emManut"]
    ws["C10"] = pa["emManut"]
    for rr in (9, 10):
        ws.cell(row=rr, column=1).font = F_BLD
        for c in (1, 2, 3):
            ws.cell(row=rr, column=c).border = BORDER

    # grafico de barras agrupado
    dados_g = [
        ("KM (km)", km_j, km_a, "#F59E0B", "#2563EB"),
        ("Litros (L)", lit_j, lit_a, "#F59E0B", "#2563EB"),
        ("CO₂ (kg)", co2_j, co2_a, "#F59E0B", "#2563EB"),
    ]
    base = 12
    for i, (rot, vj, va, cj, ca) in enumerate(dados_g):
        rr = base + i * 3
        ws.cell(row=rr, column=2, value="JULHO")
        ws.cell(row=rr, column=3, value=vj)
        ws.cell(row=rr + 1, column=2, value="AGOSTO")
        ws.cell(row=rr + 1, column=3, value=va)
        grafico = BarChart()
        grafico.type = "col"
        grafico.style = 10
        grafico.width = 7.5
        grafico.height = 5
        grafico.title = rot
        grafico.y_axis.title = rot.split(" ")[0]
        grafico.add_data(
            Reference(ws, min_col=3, min_row=rr, max_row=rr + 1),
            titles_from_data=False)
        grafico.set_categories(Reference(ws, min_col=2, min_row=rr, max_row=rr + 1))
        if i == 0:
            grafico.dataLabels = DataLabelList()
            grafico.dataLabels.showVal = True
        ws.add_chart(grafico, f"G{base + i * 3}")

    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 18
    for col in "FGHI":
        ws.column_dimensions[col].width = 14

    # =====================================================================
    # ABA 2 — COMBUSTÍVEL
    # =====================================================================
    ws2 = wb.create_sheet("Combustível")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "COMBUSTÍVEL — Diesel e Gasolina (litros e CO₂)"
    ws2["A1"].font = F_TIT
    ws2["A1"].fill = FILL_TIT
    ws2["A1"].alignment = AL_C
    ws2.row_dimensions[1].height = 30

    ws2["A3"] = "Combustível"
    ws2["B3"] = "Litros JULHO"
    ws2["C3"] = "Litros AGOSTO"
    ws2["D3"] = "Δ litros"
    ws2["E3"] = "Δ litros (%)"
    ws2["F3"] = "CO₂ JULHO (kg)"
    ws2["G3"] = "CO₂ AGOSTO (kg)"
    ws2["H3"] = "Δ CO₂ (kg)"
    estilo_cabecalho(ws2, 3, 8)

    comb = [
        ("Diesel", jul["litrosD"], ago["litrosD"], jul["diesel"], ago["diesel"]),
        ("Gasolina", jul["litrosG"], ago["litrosG"], jul["gasolina"], ago["gasolina"]),
    ]
    r = 4
    for nome, lj, la, cj, ca in comb:
        ws2.cell(row=r, column=1, value=nome).font = F_BLD
        ws2.cell(row=r, column=2, value=lj)
        ws2.cell(row=r, column=3, value=la)
        aplicar_delta(ws2.cell(row=r, column=4), la - lj)
        ws2.cell(row=r, column=5).value = pct_num((la - lj) / lj * 100)
        ws2.cell(row=r, column=5).number_format = "+0.0%;-0.0%"
        ws2.cell(row=r, column=6, value=cj)
        ws2.cell(row=r, column=7, value=ca)
        aplicar_delta(ws2.cell(row=r, column=8), ca - cj)
        for c in range(1, 9):
            ws2.cell(row=r, column=c).border = BORDER
            if c > 1:
                ws2.cell(row=r, column=c).number_format = '#,##0.0'
                ws2.cell(row=r, column=c).alignment = AL_C
        ws2.row_dimensions[r].height = 20
        r += 1

    ws2.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws2.column_dimensions[col].width = 16

    # grafico CO2 por combustivel
    ws2["A10"] = "Atual"
    ws2["B10"] = "Mês"
    ws2["C10"] = "Diesel (kg)"
    ws2["D10"] = "Gasolina (kg)"
    ws2["A11"] = "Diesel"
    ws2["A12"] = "Gasolina"
    ws2["B11"] = "JULHO"
    ws2["B12"] = "AGOSTO"
    ws2["C11"] = jul["diesel"]
    ws2["C12"] = ago["diesel"]
    ws2["D11"] = jul["gasolina"]
    ws2["D12"] = ago["gasolina"]
    g2 = BarChart()
    g2.type = "col"
    g2.style = 10
    g2.width = 12
    g2.height = 7
    g2.title = "CO₂ (kg) por combustível — JULHO × AGOSTO"
    g2.add_data(Reference(ws2, min_col=3, min_row=11, max_col=4, max_row=12),
                titles_from_data=True, from_rows=False)
    g2.set_categories(Reference(ws2, min_col=2, min_row=11, max_row=12))
    g2.dataLabels = DataLabelList()
    g2.dataLabels.showVal = True
    ws2.add_chart(g2, "G10")

    # pizza participação Diesel x Gasolina em AGOSTO
    ws2["A21"] = "Participação no CO₂ de AGOSTO"
    ws2["A22"] = "Diesel"
    ws2["B22"] = ago["diesel"]
    ws2["A23"] = "Gasolina"
    ws2["B23"] = ago["gasolina"]
    pz = PieChart()
    pz.title = "Participação no CO₂ — AGOSTO"
    pz.width = 7.5
    pz.height = 7
    pz.add_data(Reference(ws2, min_col=2, min_row=22, max_row=23), titles_from_data=False)
    pz.set_categories(Reference(ws2, min_col=1, min_row=22, max_row=23))
    pz.dataLabels = DataLabelList()
    pz.dataLabels.showPercent = True
    for i, cor in enumerate(["#F59E0B", "#EF4444"]):
        pz.series[0].data_points.append(DataPoint(idx=i, spPr=GraphicalProperties(solidFill=cor)))
    ws2.add_chart(pz, "A26")

    # =====================================================================
    # ABA 3 — VEÍCULOS
    # =====================================================================
    ws3 = wb.create_sheet("Veículos")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "VEÍCULOS — KM e CO₂ de AGOSTO (comparado a JULHO)"
    ws3["A1"].font = F_TIT
    ws3["A1"].fill = FILL_TIT
    ws3["A1"].alignment = AL_C
    ws3.row_dimensions[1].height = 30

    ws3["A3"] = "Veículo"
    ws3["B3"] = "Modelo"
    ws3["C3"] = "Combustível"
    ws3["D3"] = "Km no mês"
    ws3["E3"] = "Litros"
    ws3["F3"] = "CO₂ AGOSTO (kg)"
    ws3["G3"] = "CO₂ JULHO (kg)"
    ws3["H3"] = "Δ CO₂ (%)"
    estilo_cabecalho(ws3, 3, 8)

    jul_por_tag = {r[0]: r[5] for r in dados["consumo"]["JULHO"]}
    rows = sorted(dados["consumo"]["AGOSTO"], key=lambda r: -r[5])
    r = 4
    for tag, modelo, comb, km, litros, co2 in rows:
        ws3.cell(row=r, column=1, value=tag).font = F_BLD
        ws3.cell(row=r, column=2, value=modelo)
        ws3.cell(row=r, column=3, value=comb)
        ws3.cell(row=r, column=4, value=km)
        ws3.cell(row=r, column=5, value=litros)
        ws3.cell(row=r, column=6, value=co2)
        ant = jul_por_tag.get(tag)
        if ant:
            ws3.cell(row=r, column=7, value=ant)
            frac_pct = (co2 - ant) / ant
            ws3.cell(row=r, column=8).value = frac_pct
            ws3.cell(row=r, column=8).number_format = "+0.0%;-0.0%"
            cel = ws3.cell(row=r, column=8)
            cel.font = F_POS if frac_pct <= 0 else F_NEG
        else:
            ws3.cell(row=r, column=7, value="—")
            ws3.cell(row=r, column=8, value="—")
        for c in range(1, 9):
            ws3.cell(row=r, column=c).border = BORDER
            if c in (4, 5, 6, 7, 8):
                ws3.cell(row=r, column=c).number_format = '#,##0.0'
                ws3.cell(row=r, column=c).alignment = AL_C
        r += 1

    for col, w in zip("ABCDEFGH", (10, 18, 13, 12, 10, 15, 14, 12)):
        ws3.column_dimensions[col].width = w

    ws3["A15"] = "Atual"
    ws3["B15"] = "Veículo"
    ws3["C15"] = "CO₂ AGO (kg)"
    for idx, (tag, _, _, _, _, co2) in enumerate(rows):
        rr = 16 + idx
        ws3.cell(row=rr, column=2, value=tag)
        ws3.cell(row=rr, column=3, value=co2)
    g3 = BarChart()
    g3.type = "col"
    g3.style = 10
    g3.width = 14
    g3.height = 8
    g3.title = "Top CO₂ por veículo — AGOSTO"
    g3.add_data(Reference(ws3, min_col=3, min_row=16, max_row=16 + len(rows)),
                titles_from_data=False, from_rows=True)
    g3.set_categories(Reference(ws3, min_col=2, min_row=16, max_row=16 + len(rows)))
    ws3.add_chart(g3, "E15")

    # =====================================================================
    # ABA 4 — MESES (TENDÊNCIA)
    # =====================================================================
    ws4 = wb.create_sheet("Meses")
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "KM TOTAL POR MÊS (toda a frota)"
    ws4["A1"].font = F_TIT
    ws4["A1"].fill = FILL_TIT
    ws4["A1"].alignment = AL_C
    ws4.row_dimensions[1].height = 30

    ws4["A3"] = "Mês"
    ws4["B3"] = "KM total"
    ws4["C3"] = "Veículos"
    ws4["D3"] = "CO₂ (kg)"
    ws4["E3"] = "CO₂ Diesel+Gasolina"
    estilo_cabecalho(ws4, 3, 5)

    co2_por_mes = {c["mes"]: c for c in dados["consumoTend"]}
    r = 4
    for t in dados["tendencia"]:
        ws4.cell(row=r, column=1, value=t["mes"]).font = F_BLD
        ws4.cell(row=r, column=2, value=t["km"])
        ws4.cell(row=r, column=3, value=t["veic"])
        c = co2_por_mes.get(t["mes"])
        if c:
            ws4.cell(row=r, column=4, value=c["total"])
            ws4.cell(row=r, column=5,
                     value=f'{numerico(c["diesel"])} + {numerico(c["gasolina"])}')
        else:
            ws4.cell(row=r, column=4, value="—")
            ws4.cell(row=r, column=5, value="—")
        for cc in range(1, 6):
            ws4.cell(row=r, column=cc).border = BORDER
        r += 1

    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 24

    ws4["B16"] = "Mês"
    ws4["C16"] = "KM"
    for idx, t in enumerate(dados["tendencia"]):
        rr = 17 + idx
        ws4.cell(row=rr, column=2, value=t["mes"])
        ws4.cell(row=rr, column=3, value=t["km"])
    g4 = BarChart()
    g4.type = "col"
    g4.style = 10
    g4.width = 18
    g4.height = 8
    g4.title = "KM total por mês"
    g4.add_data(Reference(ws4, min_col=3, min_row=17, max_row=17 + len(dados["tendencia"]) - 1),
                titles_from_data=False, from_rows=True)
    g4.set_categories(Reference(ws4, min_col=2, min_row=17, max_row=17 + len(dados["tendencia"]) - 1))
    ws4.add_chart(g4, "E16")

    # =====================================================================
    # ABA 5 — TEXTOS
    # =====================================================================
    ws5 = wb.create_sheet("Textos da Apresentação")
    ws5.merge_cells("A1:B1")
    ws5["A1"] = "TEXTOS PRONTOS PARA APRESENTAÇÃO — JULHO × AGOSTO"
    ws5["A1"].font = F_TIT
    ws5["A1"].fill = FILL_TIT
    ws5["A1"].alignment = AL_C
    ws5.row_dimensions[1].height = 30

    pct_km = pct((km_a - km_j) / km_j * 100)
    pct_lit = pct((lit_a - lit_j) / lit_j * 100)
    pct_co2 = pct((co2_a - co2_j) / co2_j * 100)
    pct_med = pct((med_a - med_j) / med_j * 100)
    pct_veic = pct((veic_a - veic_j) / veic_j * 100, 0)
    pct_dl = pct((ago["litrosD"] - jul["litrosD"]) / jul["litrosD"] * 100)
    pct_gl = pct((ago["litrosG"] - jul["litrosG"]) / jul["litrosG"] * 100)
    pct_dc = pct((ago["diesel"] - jul["diesel"]) / jul["diesel"] * 100)
    pct_gc = pct((ago["gasolina"] - jul["gasolina"]) / jul["gasolina"] * 100)

    textos = [
        ("FALA DE ABERTURA (1 min)",
         f"\"Em AGOSTO a frota andou {br(km_a, 0)} km — contra {br(km_j, 0)} km em "
         f"JULHO, queda de {br(km_j - km_a, 0)} km ({pct_km}) com {veic_a} veículos "
         f"ativos (JULHO teve {veic_j}). Essa redução derrubou o consumo de combustível "
         f"e, direto, as emissões de CO₂. Menos rodagem, menos custo, menos emissão, "
         f"menos exposição dos motoristas na estrada.\""),
        ("SLIDE 1 — Quadro comparativo",
         f"KM total: {br(km_j, 0)} km → {br(km_a, 0)} km (Δ {d(km_a - km_j, 0)} km, "
         f"{pct_km})  •  Veículos: {veic_j} → {veic_a} ({pct_veic})  •  Litros totais: "
         f"{br(lit_j, 2)} L → {br(lit_a, 2)} L (Δ {d(lit_a - lit_j, 2)} L, {pct_lit})  •  "
         f"CO₂ total: {br(co2_j, 2)} kg → {br(co2_a, 2)} kg (Δ {d(co2_a - co2_j, 2)} kg, "
         f"{pct_co2})  •  Média por veículo: {br(med_j, 1)} km → {br(med_a, 1)} km."),
        ("SLIDE 2 — Quilometragem (o que caiu)",
         f"Média por veículo caiu de {br(med_j, 1)} km para {br(med_a, 1)} km "
         f"(≈ {br(abs((med_a - med_j) / med_j * 100), 0)}%). {ts} foi o que mais rodou "
         f"em AGOSTO: {br(ts_ka, 0)} km ({br(ts_kj, 0)} km em JULHO). TN 01 (Pulse) "
         f"subiu: {br(km01_a, 0)} km contra {br(km01_j, 0)} km.{tn76_txt} Frota ficou "
         f"{pa['pctOp']:.0f}% operacional em AGOSTO ({pa['ativos']} de {veic_a}), contra "
         f"{pj['pctOp']:.0f}% em JULHO."),
        ("SLIDE 3 — Combustível (Diesel cai forte, Gasolina sobe)",
         f"DIESEL: {br(jul['litrosD'], 2)} L → {br(ago['litrosD'], 2)} L "
         f"(Δ {d(ago['litrosD'] - jul['litrosD'], 2)} L, {pct_dl}) e CO₂ "
         f"{br(jul['diesel'], 2)} kg → {br(ago['diesel'], 2)} kg "
         f"(Δ {d(ago['diesel'] - jul['diesel'], 2)} kg, {pct_dc}).  GASOLINA: "
         f"{br(jul['litrosG'], 2)} L → {br(ago['litrosG'], 2)} L "
         f"(Δ {d(ago['litrosG'] - jul['litrosG'], 2)} L, {pct_gl}) e CO₂ "
         f"{br(jul['gasolina'], 2)} kg → {br(ago['gasolina'], 2)} kg "
         f"(Δ {d(ago['gasolina'] - jul['gasolina'], 2)} kg, {pct_gc}). Participação: Diesel "
         f"de {p_diesel_j:.0f}% para {p_diesel_a:.0f}% das emissões; Gasolina de "
         f"{100 - p_diesel_j:.0f}% para {100 - p_diesel_a:.0f}%. {aumento_txt}"),
        ("SLIDE 4 — Emissões de CO₂",
         f"CO₂ evitado em AGOSTO: ~{br((co2_j - co2_a), 1)} kg (~{br((co2_j - co2_a) / 1000, 2)} "
         f"tonelada vs JULHO, {pct_co2}). Top emissores de AGOSTO: {top_txt}. "
         f"{top0[0] if top0 else ''} sozinha ≈ {br(pct_top, 0)}% do CO₂ do mês "
         + (f"e, ainda assim, emitiu ~{br(diff_top, 0)} kg a menos que em JULHO."
            if diff_top and diff_top > 0 else "e segue como principal emissora.")),
        ("SLIDE 5 — Benefícios da redução",
         f"1) Custo de combustível: {d(lit_a - lit_j, 0)} litros no mês. "
         f"2) Emissões: {d(co2_a - co2_j, 2)} kg CO₂e, apoia metas ESG. "
         f"3) Desgaste da frota: menos km = menos manutenção, pneus e revisões. "
         f"4) Operação: frota {pa['pctOp']:.0f}% operacional em AGOSTO"
         + (", sem veículos em manutenção nem mobilização."
            if pa["emManut"] == 0 else f", com {pa['emManut']} veículo(s) em manutenção.")),
        ("SLIDE 6 — Exposição dos motoristas",
         f"{d(km_a - km_j, 0)} km rodados significa centenas de horas a menos ao volante "
         f"no mês: menos exposição a acidentes, estradas precárias e fadiga. Motoristas de "
         f"caminhonete (frota Diesel) foram os mais beneficiados — concentraram "
         f"{br(pct_km_diesel, 0)}%+ do km e tiveram as maiores reduções (ex.: "
         f"{ex_reducoes}). Dias de operação mais leves → turnos mais curtos, mais "
         f"segurança e qualidade de vida."),
        ("FALA DE ENCERRAMENTO",
         f"\"AGOSTO entregou menos km ({pct_km}), menos combustível ({pct_lit}) e menos "
         f"CO₂ ({pct_co2}), mantendo a frota {pa['pctOp']:.0f}% operacional. O único aumento "
         f"foi a gasolina do veículo leve ({pct_gc}), irrelevante no total. É o retrato de "
         f"uma operação mais enxuta, mais segura e mais sustentável.\""),
    ]
    ws5["A3"] = "Seção"
    ws5["B3"] = "Texto"
    estilo_cabecalho(ws5, 3, 2)
    r = 4
    for titulo, corpo in textos:
        ws5.cell(row=r, column=1, value=titulo).font = F_BLD
        ws5.cell(row=r, column=2, value=corpo).alignment = AL_L
        ws5.row_dimensions[r].height = 66
        for c in (1, 2):
            ws5.cell(row=r, column=c).border = BORDER
        ws5.cell(row=r, column=1).alignment = AL_L
        r += 1

    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["B"].width = 110

    wb.save(SAIDA)
    print(f"Gerado: {SAIDA} ({SAIDA.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()