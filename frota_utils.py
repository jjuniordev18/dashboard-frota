"""
frota_utils.py — ETL corrigido + corretor da planilha de KM da frota.

Lógica de negócio criada a partir da planilha real:
  - A aba mensal tem 3 blocos no mesmo sheet:
      1. VEÍCULOS CADASTRADOS: TAG, KM INICIAL e leituras acumuladas S1..S5 do hodômetro.
      2. Tabela antiga "SEMANA 02 / KM TOTAL" (restos de um template antigo — ignorada).
      3. ACOMPANHAMENTO KM SEMANA: TAG + semanas S1..S5 + coluna ACUM.
  - O bug: as células da coluna ACUM. usam =SUM(B..F) de fórmulas de diferença que
    quebram quando há leitura vazia/texto (NaN -> erro), referência errada (=K10-J9)
    ou #REF! (texto em KM INICIAL). Resultado: ACUM. fica negativo/lixo e o dashboard
    mostra 0 para o mês de AGOSTO.
  - Correção: o KM mensal é calculado pelas leituras do hodômetro do bloco de
    cadastro:  KM_MES = última leitura válida - KM INICIAL (ou soma dos incrementos
    positivos quando não há KM INICIAL).
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

# ----------------------------------------------------------------------------
# Ordem cronológica dos meses (a planilha usa nomes soltos a partir de Janeiro)
# ----------------------------------------------------------------------------
MESES_ORDEM = [
    "OUTUBRO 2025", "NOVEMBRO 2025", "DEZEMBRO 2025",
    "JANEIRO", "FEVEREIRO", "MAR\u00c7O", "ABRIL", "MAIO",
    "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO",
    "NOVEMBRO", "DEZEMBRO",
]

ALIAS_MES = {
    "MAR\u00c7O": "MAR\u00c7O", "MARCO": "MAR\u00c7O",
    "ABRIL": "ABRIL", "MAIO": "MAIO",
    "OUTUBRO 2025": "OUTUBRO 2025", "NOVEMBRO 2025": "NOVEMBRO 2025",
    "DEZEMBRO 2025": "DEZEMBRO 2025",
    "JANEIRO": "JANEIRO", "FEVEREIRO": "FEVEREIRO",
    "JUNHO": "JUNHO", "JULHO": "JULHO", "AGOSTO": "AGOSTO",
    "SETEMBRO": "SETEMBRO", "OUTUBRO": "OUTUBRO",
    "NOVEMBRO": "NOVEMBRO", "DEZEMBRO": "DEZEMBRO",
}

MES_NOMES = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MAR\u00c7O", 4: "ABRIL", 5: "MAIO", 6: "JUNHO",
    7: "JULHO", 8: "AGOSTO", 9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO",
    12: "DEZEMBRO",
}


def _rotulo_mes(dt) -> str:
    """Rótulo do mês no padrão do projeto: sem ano p/ 2026, com ano caso contrário."""
    nome = MES_NOMES.get(dt.month, str(dt.month).zfill(2))
    return nome if dt.year == 2026 else f"{nome} {dt.year}"


def chave_frota(mes_ref) -> str:
    """Chave de mês da frota p/ juntar com o CO2: 'AGOSTO' -> 'AGOSTO 2026'."""
    s = str(mes_ref).strip()
    if re.search(r"\d{4}$", s):
        return s
    return s + " 2026"


def normalizar_mes(nome) -> str:
    """Normaliza o nome da aba (espaços extras / acentuação) e devolve o mês canônico."""
    limpo = " ".join(str(nome).strip().split()).upper()
    limpo = limpo.replace("\u00c0", "A")
    return ALIAS_MES.get(limpo, limpo.title())


def ordem_mes(nome_canonico) -> int:
    if nome_canonico in MESES_ORDEM:
        return MESES_ORDEM.index(nome_canonico)
    return 1000


# ----------------------------------------------------------------------------
# Conversão segura de número (pula textos como 'Não informado', 'Manutenção'...)
# ----------------------------------------------------------------------------
TOKENS_INVALIDOS = [
    "ERROR", "#", "MANUTEN", "MATUTEN", "MOBILIZ", "DESMOBILIZ", "LAVADOR",
    "N\u00c3O", "N\u00c3OINFORMADO", "PENDENTE", "OK", "EMDIA", "EM DIAS",
    "INFORMADO", "REF!", "VALUE!", "AGUARDANDO", "LAVAGEM", "PROGR",
]


def safe_num(val):
    """Converte para float; retorna NaN para texto inválido / vazio."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return np.nan
    if isinstance(val, bool):
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().upper().replace("\u00a0", " ")
    s = re.sub(r"\s+", "", s)
    if s in ("", "NAN", "NONE", "N/D"):
        return np.nan
    if any(t in s for t in TOKENS_INVALIDOS):
        return np.nan
    try:
        return float(s.replace(".", "").replace(",", "."))
    except Exception:
        return np.nan


def eh_float(v):
    return isinstance(v, (int, float)) and not (isinstance(v, float) and np.isnan(v))


def canonical_tag(tag) -> str:
    """Padroniza a TAG (a planilha alterna 'TN78' e 'TN 78')."""
    t = re.sub(r"\s+", "", str(tag)).upper()
    m = re.fullmatch(r"TN(\d+)", t)
    if m:
        return f"TN {m.group(1)}"
    return t


# ----------------------------------------------------------------------------
# Extração de UMA aba mensal (bloco de cadastro = fonte confiável do KM)
# ----------------------------------------------------------------------------
def extrair_aba_mensal(df_raw: pd.DataFrame):
    """Retorna DataFrame por veículo da aba mensal, ou None se a aba não for mensal."""
    rows = df_raw.values

    # 1) localiza a linha de cabeçalho que contém 'KM INICIAL'
    head_idx = None
    for i in range(min(len(rows), 40)):
        row = [str(c).strip().upper() for c in rows[i]]
        if "KM INICIAL" in row and any(re.match(r"^S\d+$", c) for c in row):
            head_idx = i
            break
    if head_idx is None:
        return None

    header = [str(c).strip().upper() for c in rows[head_idx]]

    def col(token):
        for j, h in enumerate(header):
            if h == token:
                return j
        return None

    tag_j = col("TAG")
    ini_j = col("KM INICIAL")
    tipo_j = col("TIPO")
    modelo_j = col("MODELOS")
    placa_j = col("PLACA")
    locadora_j = col("LOCADORA")
    localidade_j = col("LOCALIDADE")

    s_idx = {k: col(f"S{k}") for k in range(1, 6) if col(f"S{k}") is not None}
    if tag_j is None or not s_idx:
        return None

    registros = []
    for i in range(head_idx + 1, len(rows)):
        row = rows[i]
        # fim do bloco de cadastro: a aba empilha outras tabelas depois
        if not pd.isna(row[0]):
            a0 = str(row[0]).strip().upper()
            if a0.startswith(("VE\u00cdCULOS", "VEICULOS", "ACOMPANHAMENTO")) \
                    or a0 in ("TAG", "TIPO"):
                break
        if any("KM TOTAL" in str(x).upper() for x in row if not pd.isna(x)):
            break
        if pd.isna(row[tag_j]):
            continue
        tag = str(row[tag_j]).strip()
        if not tag or tag.upper() in ("TAG", "VE\u00cdCULOS", "VEICULOS", "ACOMPANHAMENTO"):
            continue
        # a linha pode ser repetição de cabeçalho em tabelas empilhadas
        if tag.upper() == "KM SEMANA":
            continue

        ini = safe_num(row[ini_j]) if ini_j is not None else np.nan
        leituras = {k: safe_num(row[j]) for k, j in s_idx.items()}

        # --- status operacional a partir dos textos nas células de semana ---
        status = "Operacional"
        for j in s_idx.values():
            s = str(row[j]).upper()
            if any(t in s for t in ("MANUTEN", "MATUTEN")):
                status = "Manuten\u00e7\u00e3o"
                break
            if any(t in s for t in ("MOBILIZ", "DESMOBILIZ")):
                status = "Mobiliza\u00e7\u00e3o"
                break
            if any(t in s for t in ("AGUARDANDO", "LAVADOR")):
                status = "Pendente"
                break
            if any(t in s for t in ("ERROR", "#", "REF!", "VALUE!")):
                status = "Erro Dados"
                break
            if "PENDENTE" in s:
                status = "Pendente"
                break

        # --- KM do mês: última leitura - KM INICIAL ---
        vals = [leituras[k] for k in s_idx.keys()]
        vals_pos = [v for v in vals if eh_float(v) and v > 0]
        # última leitura válida (por ordem de semana) — a mais confiável,
        # pois o hodômetro pode ter sido trocado/reiniciado no meio do mês
        final = vals_pos[-1] if vals_pos else np.nan

        km_mes = np.nan
        if eh_float(ini) and ini > 0 and eh_float(final) and final > ini:
            km_mes = final - ini
        if not (eh_float(km_mes) and 0 < km_mes <= 30000):
            # fallback: soma de incrementos positivos (quando não há KM INICIAL
            # ou o KM INICIAL está corrompido)
            incr = []
            prev = None
            for k in sorted(s_idx.keys()):
                v = leituras[k]
                if not (eh_float(v) and v > 0):
                    continue
                if prev is not None and v > prev and (v - prev) <= 30000:
                    incr.append(v - prev)
                prev = v
            km_mes = float(sum(incr)) if incr else np.nan

        # --- KM semanal (incrementos por semana) p/ o gráfico de evolução ---
        semana = {k: np.nan for k in range(1, 6)}
        prev = ini if (eh_float(ini) and ini > 0) else None
        for k in sorted(s_idx.keys()):
            v = leituras[k]
            if not (eh_float(v) and v > 0):
                continue
            if prev is not None and v > prev and (v - prev) <= 30000:
                semana[k] = v - prev
            prev = v

        registro = {
            "MES_REF": None,
            "TAG": canonical_tag(tag),
            "TIPO": str(row[tipo_j]).strip() if tipo_j is not None and not pd.isna(row[tipo_j]) else "",
            "MODELO": str(row[modelo_j]).strip() if modelo_j is not None and not pd.isna(row[modelo_j]) else "",
            "PLACA": str(row[placa_j]).strip() if placa_j is not None and not pd.isna(row[placa_j]) else "",
            "LOCADORA": str(row[locadora_j]).strip() if locadora_j is not None and not pd.isna(row[locadora_j]) else "",
            "LOCALIDADE": str(row[localidade_j]).strip() if localidade_j is not None and not pd.isna(row[localidade_j]) else "",
            "KM_INICIAL": ini,
            "KM_FINAL": final,
            "KM_MES": km_mes,
            "STATUS": status,
            **{f"S{k}": leituras.get(k, np.nan) for k in range(1, 6)},
            **{f"W{k}": semana[k] for k in range(1, 6)},
        }
        registros.append(registro)

    if not registros:
        return None
    df_mes = pd.DataFrame(registros)
    df_mes["KM_MES"] = pd.to_numeric(df_mes["KM_MES"], errors="coerce")
    return df_mes


# ----------------------------------------------------------------------------
# Processamento completo do arquivo (todas as abas mensais)
# ----------------------------------------------------------------------------
def processar_planilha(arquivo) -> pd.DataFrame:
    xls = pd.ExcelFile(arquivo)
    frames = []
    for sheet in xls.sheet_names:
        df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        df_mes = extrair_aba_mensal(df_raw)
        if df_mes is None:
            continue
        canon = normalizar_mes(sheet)
        df_mes["MES_REF"] = canon
        df_mes["MES_ORDEM"] = ordem_mes(canon)
        frames.append(df_mes)

    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)
    # ordena por mês e depois por TAG
    df = df.sort_values(["MES_ORDEM", "TAG"]).reset_index(drop=True)

    # inferência de Tipo quando o bloco não tem a coluna TIPO
    if df["TIPO"].isna().all() or (df["TIPO"] == "").all():
        df["TIPO"] = np.where(df["TAG"].str.upper().str.startswith("TN 01"), "Leve", "Caminhonete")

    status_validos = ["Manuten\u00e7\u00e3o", "Mobiliza\u00e7\u00e3o", "Operacional",
                      "Pendente", "Lavador", "Erro Dados"]
    df.loc[~df["STATUS"].isin(status_validos), "STATUS"] = "Operacional"
    return df


# ----------------------------------------------------------------------------
# CO2 / ESG — Dashboard_consumo_co2.xlsx (fatores + registro mensal)
# ----------------------------------------------------------------------------
def _achar_aba(wb, pistas) -> str | None:
    for s in wb.sheetnames:
        chave = "".join(ch for ch in s if ch.isalnum()).upper()
        if any(p in chave for p in pistas):
            return s
    return None


def _linha_cabecalho(df_raw, chaves) -> int | None:
    for i, row in df_raw.iterrows():
        cel = " ".join(str(v) for v in row.dropna()).upper()
        if all(c in cel for c in chaves):
            return i
    return None


def _col_por_pista(hdr, pistao) -> int | None:
    for i, h in enumerate(hdr):
        hs = str(h).upper()
        if pistao in hs or hs in pistao:
            return i
    return None


def processar_co2(arquivo, so_frota=True) -> pd.DataFrame:
    """Lê 'Dashboard_consumo_co2.xlsx': fatores de emissão + registro mensal de
    consumo. Recalcula CO2E = quantidade x fator e devolve o DataFrame limpo
    (colunas: MES_REF, MES_CHAVE, MES_DT, FONTE, ATIVIDADE, USO, QUANTIDADE,
    UNIDADE, FATOR, CO2E).

    `so_frota=True` (padrão): mantém apenas os registros de VEÍCULOS (atividade
    'Frota...') — lançamentos de geradores/energia são desconsiderados.
    """
    if hasattr(arquivo, "read"):
        arquivo.seek(0)
        payload = arquivo.read()
    elif isinstance(arquivo, (str, Path)):
        payload = Path(arquivo).read_bytes()
    else:
        payload = arquivo

    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    aba_f = _achar_aba(wb, ["FATOR"])
    aba_r = _achar_aba(wb, ["REGISTRO"])
    if aba_f is None or aba_r is None:
        return pd.DataFrame()

    # --- fatores de emissão: FONTE | UNIDADE | FATOR | REFERÊNCIA ---
    raw_f = pd.read_excel(io.BytesIO(payload), sheet_name=aba_f, header=None, dtype=object)
    hi_f = _linha_cabecalho(raw_f, ["FONTE", "UNIDADE", "FATOR"])
    if hi_f is None:
        return pd.DataFrame()
    fdf = raw_f.iloc[hi_f + 1:].copy()
    fdf.columns = [str(c) for c in raw_f.iloc[hi_f]]

    fatores = {}
    for _, r in fdf.iterrows():
        nome = str(r.iloc[0]).strip()
        token = re.split(r"[\(/\s]", nome)[0].strip().upper()
        fatores[token] = {
            "NOME": nome,
            "FATOR": safe_num(r.iloc[2]),
            "UNIDADE": str(r.iloc[1]).strip(),
        }

    # --- registro mensal de consumo ---
    raw_r = pd.read_excel(io.BytesIO(payload), sheet_name=aba_r, header=None, dtype=object)
    hr = _linha_cabecalho(raw_r, ["QUANTIDADE", "ATIVIDADE"])
    if hr is None:
        return pd.DataFrame()
    rd = raw_r.iloc[hr + 1:].copy()
    hdr = [str(c) for c in raw_r.iloc[hr]]

    ci_mes = _col_por_pista(hdr, "M\u00caS")
    if ci_mes is None:
        ci_mes = _col_por_pista(hdr, "MES")
    ci_fonte = _col_por_pista(hdr, "FONTE")
    ci_ativ = _col_por_pista(hdr, "ATIVIDADE")
    ci_qtd = _col_por_pista(hdr, "QUANTIDADE")
    ci_un = _col_por_pista(hdr, "UNIDADE")
    ci_emi = _col_por_pista(hdr, "EMISS\u00c3O")

    registros = []
    for _, row in rd.iterrows():
        mes_val = row.iloc[ci_mes] if ci_mes is not None else None
        mes_dt = pd.to_datetime(mes_val, errors="coerce")
        if pd.isna(mes_dt):
            continue

        fonte = str(row.iloc[ci_fonte]).strip() if ci_fonte is not None else ""
        info = fatores.get(fonte.upper())
        fator = (info["FATOR"] if info else np.nan)
        unid = (info["UNIDADE"] if info else
                (str(row.iloc[ci_un]).strip() if ci_un is not None else ""))

        qtd = safe_num(row.iloc[ci_qtd]) if ci_qtd is not None else np.nan
        co2 = (qtd * fator) if (eh_float(qtd) and eh_float(fator)) else np.nan
        if np.isnan(co2) and ci_emi is not None:
            co2 = safe_num(row.iloc[ci_emi])

        ativ = str(row.iloc[ci_ativ]).strip() if ci_ativ is not None else ""
        uso = "Frota" if "FROTA" in ativ.upper() else "Infra"

        registros.append({
            "MES_REF": _rotulo_mes(mes_dt),
            "MES_CHAVE": f"{MES_NOMES.get(mes_dt.month, mes_dt.month)} {mes_dt.year}",
            "MES_DT": mes_dt,
            "FONTE": info["NOME"] if info else fonte,
            "FONTE_CURTA": fonte,
            "ATIVIDADE": ativ,
            "USO": uso,
            "QUANTIDADE": qtd,
            "UNIDADE": unid,
            "FATOR": fator,
            "CO2E": co2,
        })

    if not registros:
        return pd.DataFrame()
    df = pd.DataFrame(registros)
    if so_frota:
        df = df[df["USO"] == "Frota"].copy()
    df["MES_ORDEM"] = df["MES_DT"].apply(lambda d: d.toordinal())
    df = df.sort_values(["MES_DT", "FONTE_CURTA"]).reset_index(drop=True)
    return df


def resumo_co2(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega o CO2 por mês (total + por fonte + por uso), em ordem cronológica."""
    if df.empty:
        return pd.DataFrame()
    ag = (df.groupby("MES_REF", as_index=False)
          .agg(MES_ORDEM=("MES_ORDEM", "first"),
               CO2_TOTAL=("CO2E", "sum"),
               FROTA=("USO", lambda s: float(df.loc[s.index[s == "Frota"].tolist(), "CO2E"].sum())),
               INFRA=("USO", lambda s: float(df.loc[s.index[s == "Infra"].tolist(), "CO2E"].sum()))))
    ag = ag.sort_values("MES_ORDEM").reset_index(drop=True)
    return ag


# ----------------------------------------------------------------------------
# CONSUMO DE CO2 DA FROTA (km real -> litros estimados -> kgCO2e)
# Combustível por veículo: Leve (Nivus/Pulse/Tracker) = Gasolina; Caminhonetes
# (Hilux, Ranger, S10, Frontier, Strada) = Diesel.
# ----------------------------------------------------------------------------
FATOR_CO2 = {"Gasolina": 2.16, "Diesel": 2.68}

GASOLINA_DEFAULT = 11.0
DIESEL_DEFAULT = 9.0

CONSUMO_KM_POR_L = {
    "Gasolina": {
        "NIVUS": 11.0, "PULSE": 11.5, "TRACKER": 11.0,
    },
    "Diesel": {
        "HILUX": 9.5, "RANGER": 9.5, "S10": 9.5,
        "FRONTIER": 9.0, "STRADA": 10.0,
    },
}

_MODELOS_GASOLINA = {"NIVUS", "PULSE", "TRACKER"}


def classificar_combustivel(tipo, modelo) -> str:
    """Gasolina só para o veículo leve (TN 01: Tracker/Nivus/Pulse); o resto é
    Caminhonete e roda a Diesel."""
    m = re.sub(r"\s+", "", str(modelo)).upper().replace("HILLUX", "HILUX")
    if str(tipo).strip().title() == "Leve" or m in _MODELOS_GASOLINA:
        return "Gasolina"
    return "Diesel"


def km_por_litro_modelo(tipo, modelo) -> float:
    """Consumo médio assumido (km/L) do modelo — mistura cidade/estrada."""
    comb = classificar_combustivel(tipo, modelo)
    m = re.sub(r"\s+", "", str(modelo)).upper().replace("HILLUX", "HILUX")
    tabela = CONSUMO_KM_POR_L[comb]
    for chave, km in tabela.items():
        if chave in m or m in chave:
            return km
    return GASOLINA_DEFAULT if comb == "Gasolina" else DIESEL_DEFAULT


def enriquecer_consumo(df: pd.DataFrame) -> pd.DataFrame:
    """Estima litros e CO2 de cada viagem/veículo-mês a partir do KM_MES real:
    LITROS = KM_MES / km_por_litro  |  CO2 = LITROS x FATOR (MCTI)."""
    d = df.copy()
    d["KM_MES"] = pd.to_numeric(d["KM_MES"], errors="coerce")
    comb = [classificar_combustivel(t, m) for t, m in zip(d["TIPO"], d["MODELO"])]
    kml = [km_por_litro_modelo(t, m) for t, m in zip(d["TIPO"], d["MODELO"])]
    d["COMBUSTIVEL"] = comb
    d["KM_POR_L"] = kml
    d["LITROS"] = d["KM_MES"] / d["KM_POR_L"]
    d["FATOR_CO2"] = d["COMBUSTIVEL"].map(FATOR_CO2)
    d["CO2E"] = d["LITROS"] * d["FATOR_CO2"]
    d.loc[d["KM_MES"].isna(), ["LITROS", "CO2E"]] = np.nan
    return d


def consolidado_consumo(d: pd.DataFrame) -> pd.DataFrame:
    """Consolidado mensal (a partir de OUTUBRO 2025): km, litros e CO2 por
    combustível, veículos e variação % vs mês anterior."""
    if d.empty:
        return pd.DataFrame()
    g = (d.groupby(["MES_ORDEM", "MES_REF"], as_index=False)
         .agg(VEICULOS=("TAG", "nunique"),
              KM_TOTAL=("KM_MES", "sum"),
              CO2_TOTAL=("CO2E", "sum")))
    sidx = pd.MultiIndex.from_arrays([g["MES_ORDEM"], g["MES_REF"]])
    l_d = d.loc[d["COMBUSTIVEL"] == "Diesel", "LITROS"] \
        .groupby([d["MES_ORDEM"], d["MES_REF"]]).sum()
    l_g = d.loc[d["COMBUSTIVEL"] == "Gasolina", "LITROS"] \
        .groupby([d["MES_ORDEM"], d["MES_REF"]]).sum()
    c_d = d.loc[d["COMBUSTIVEL"] == "Diesel", "CO2E"] \
        .groupby([d["MES_ORDEM"], d["MES_REF"]]).sum()
    c_g = d.loc[d["COMBUSTIVEL"] == "Gasolina", "CO2E"] \
        .groupby([d["MES_ORDEM"], d["MES_REF"]]).sum()
    g["LITROS_DIESEL"] = l_d.reindex(sidx).fillna(0.0).values
    g["LITROS_GASOLINA"] = l_g.reindex(sidx).fillna(0.0).values
    g["CO2_DIESEL"] = c_d.reindex(sidx).fillna(0.0).values
    g["CO2_GASOLINA"] = c_g.reindex(sidx).fillna(0.0).values
    g = g.sort_values("MES_ORDEM").reset_index(drop=True)
    g["DELTA_PCT"] = g["CO2_TOTAL"].pct_change() * 100
    return g


def co2_por_veiculo(d: pd.DataFrame) -> pd.DataFrame:
    """KM, litros e CO2 por veículo/mês (para comparar a evolução veículo a veículo)."""
    if d.empty:
        return pd.DataFrame()
    g = (d.groupby(["MES_ORDEM", "MES_REF", "TAG", "TIPO", "MODELO",
                    "COMBUSTIVEL"], as_index=False)
         .agg(KM_MES=("KM_MES", "sum"), LITROS=("LITROS", "sum"),
              CO2E=("CO2E", "sum")))
    g = g.sort_values(["MES_ORDEM", "CO2E"], ascending=[True, False]).reset_index(drop=True)
    return g


# ----------------------------------------------------------------------------
# CORRETOR: corrige a planilha (.xlsx) — conserta a coluna ACUM. e as semanas
# do bloco ACOMPANHAMENTO de TODAS as abas mensais (incl. AGOSTO).
# ----------------------------------------------------------------------------
def _corrigir_aba(ws):
    """Corrige UMA aba mensal. Retorna True se algo foi corrigido."""
    corrigi = False
    max_row = min(ws.max_row, 200)
    max_col = ws.max_column

    cadastro_header = None
    ini_j = None
    tag_j = None
    s_j = {}
    for r in range(1, max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        upper = [str(x).strip().upper() if x is not None else "" for x in cells]
        if "KM INICIAL" in upper and any(re.match(r"^S\d+$", x) for x in upper):
            cadastro_header = r
            tag_j = upper.index("TAG") if "TAG" in upper else None
            ini_j = upper.index("KM INICIAL")
            for k in range(1, 6):
                token = f"S{k}"
                if token in upper:
                    s_j[k] = upper.index(token)
            break
    if cadastro_header is None or tag_j is None or not s_j:
        return False

    # --- mapa tag -> (km_inicial, leituras) do bloco de cadastro ---
    def _fim_de_tabela(ws, r):
        """True quando a linha r inicia uma outra tabela dentro da aba."""
        a = ws.cell(row=r, column=1).value
        if a is not None:
            a = str(a).strip().upper()
            if a.startswith(("VE\u00cdCULOS", "VEICULOS", "ACOMPANHAMENTO")) \
                    or a in ("TAG", "TIPO"):
                return True
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().upper()
            if s in ("KM TOTAL", "SEMANA 02", "SEMANA 01", "SEMANA 04",
                     "ACUM.", "S1", "TAG", "TIPO"):
                return True
        return False

    mapa = {}
    for r in range(cadastro_header + 1, max_row + 1):
        if _fim_de_tabela(ws, r):
            break
        tag = ws.cell(row=r, column=tag_j + 1).value
        if tag is None:
            continue
        tag = str(tag).strip()
        if not tag or tag.upper() in ("TAG", "VE\u00cdCULOS", "VEICULOS"):
            continue
        ini = safe_num(ws.cell(row=r, column=ini_j + 1).value)
        leit = {k: safe_num(ws.cell(row=r, column=j + 1).value) for k, j in s_j.items()}
        mapa[canonical_tag(tag)] = (ini, leit)

    # --- acha o bloco ACOMPANHAMENTO (cabeçalho 'S1..S5' + 'ACUM.') ---
    acomp_row = None
    for r in range(1, max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().upper().startswith("ACOMPANHAMENTO"):
            acomp_row = r
            break
    if acomp_row is None:
        return False

    header_row = acomp_row + 1
    head = [str(ws.cell(row=header_row, column=c).value).strip().upper()
            for c in range(1, max_col + 1)]
    acum_j = None
    for c, h in enumerate(head):
        if "ACUM" in h:
            acum_j = c + 1  # coluna 1-based
            break
    if acum_j is None:
        return False

    # colunas 1-based das semanas no bloco de acompanhamento
    week_cols = {}
    for k in range(1, 6):
        for c, h in enumerate(head):
            if h == f"S{k}":
                week_cols[k] = c + 1
                break

    for r in range(header_row + 1, max_row + 1):
        tag = ws.cell(row=r, column=1).value
        if tag is None:
            continue
        tag = str(tag).strip()
        if not tag or tag.upper().startswith("ACOMPANHAMENTO"):
            break
        chave = canonical_tag(tag)
        if chave not in mapa:
            continue

        ini, leit = mapa[chave]

        # incrementos semanais (escrita progressiva do hodômetro)
        deltas = {}
        prev = ini if (eh_float(ini) and ini > 0) else None
        first_valid = None
        last_valid = None
        for k in sorted(leit.keys()):
            v = leit[k]
            if not (eh_float(v) and v > 0):
                continue
            if first_valid is None:
                first_valid = v
            if prev is not None and v > prev and (v - prev) <= 30000:
                deltas[k] = v - prev
            prev = v
            last_valid = v

        # ACUM = último hodômetro - linha de base (KM INICIAL ou 1ª leitura)
        base = ini if (eh_float(ini) and ini > 0) else first_valid
        acum = (last_valid - base) if (eh_float(base) and eh_float(last_valid)) else None
        if acum is not None and acum < 0:
            acum = sum(deltas.values()) or None
        if acum is not None and not 0 <= acum <= 30000:
            acum = sum(deltas.values()) or None

        corrigi = True
        for k in range(1, 6):
            col = week_cols.get(k)
            if col is None:
                continue
            cell = ws.cell(row=r, column=col)
            novo = deltas.get(k)
            if novo is not None:
                cell.value = float(novo)
            else:
                cell.value = None  # limpa fórmula quebrada / marcador negativo
        if acum is not None:
            ws.cell(row=r, column=acum_j).value = float(acum)
        else:
            ws.cell(row=r, column=acum_j).value = None
    return corrigi


def corrigir_workbook(arquivo_bytes: bytes):
    """Devolve (bytes_da_planilha_corrigida, lista_de_abas_corrigidas)."""
    wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=False)
    corrigidas = []
    for name in wb.sheetnames:
        try:
            if _corrigir_aba(wb[name]):
                corrigidas.append(name)
        except Exception:
            continue
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), corrigidas


# ----------------------------------------------------------------------------
# Dados processados -> resumo por mês (p/ a visão de tendência)
# ----------------------------------------------------------------------------
def resumo_mensal(df: pd.DataFrame):
    df = df.copy()
    df["KM_MES"] = pd.to_numeric(df["KM_MES"], errors="coerce").fillna(0)
    resumo = (
        df.groupby("MES_REF")
        .agg(KM_TOTAL=("KM_MES", "sum"),
             VEICULOS=("TAG", "nunique"),
             MEDIA=("KM_MES", "mean"),
             MANUTENCAO=("STATUS", lambda s: int((s == "Manuten\u00e7\u00e3o").sum())))
        .reset_index()
    )
    resumo["MES_ORDEM"] = resumo["MES_REF"].map(ordem_mes)
    resumo = resumo.sort_values("MES_ORDEM").reset_index(drop=True)
    return resumo


def timestamp_humano() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M")