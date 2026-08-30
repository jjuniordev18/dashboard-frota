"""gera planilha de apresentacao JULHO x AGOSTO com dados extraidos da landing page
(dashboard_frota.html). Cria 4 abas: Resumo, Combustivel, Veiculos e Textos.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DIR = Path(__file__).resolve().parent
HTML = DIR / "dashboard_frota.html"
SAIDA = DIR / "Apresentacao_Julho_Agosto.xlsx"

AZUL = "2563EB"
AMARELO = "F59E0B"
VERDE = "10B981"
VERMELHO = "EF4444"
CINZA = "64748B"
BANCO = "111C30"

THIN = Side(style="thin", color="D1D9E4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TIT = Font(bold=True, size=16, color="FFFFFF")
F_SEG = Font(bold=True, size=13, color="22303F")
F_HED = Font(bold=True, size=11, color="FFFFFF")
F_BLD = Font(bold=True, color="22303F")
F_POS = Font(color=VERDE, bold=True)
F_NEG = Font(color=VERMELHO, bold=True)
FILL_TIT = PatternFill("solid", fgColor=AZUL)
FILL_HED = PatternFill("solid", fgColor="14507A")
FILL_ZEB = PatternFill("solid", fgColor="F1F5FB")
AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def extrair_dados() -> dict:
    html = HTML.read_text(encoding="utf-8")
    m = re.search(r"var DADOS = (\{.*?\});\s*var PALETA", html, re.S)
    if not m:
        raise SystemExit("Não encontrei os dados no HTML.")
    return json.loads(m.group(1))


def numerico(v) -> str:
    if v is None:
        return "—"
    return str(v).replace(".", ",")


def f_km(v) -> str:
    return numerico(v) + " km"


def estilo_cabecalho(ws, linha: int, n_col: int):
    for c in range(1, n_col + 1):
        cel = ws.cell(row=linha, column=c)
        cel.font = F_HED
        cel.fill = FILL_HED
        cel.alignment = AL_C
        cel.border = BORDER


def zebrar(ws, r_inicio: int, r_fim: int, n_col: int):
    for r in range(r_inicio, r_fim + 1):
        if (r - r_inicio) % 2 == 1:
            for c in range(1, n_col + 1):
                ws.cell(row=r, column=c).fill = FILL_ZEB


def linha_total(ws, r: int, n_col: int, label: str = "TOTAL"):
    for c in range(1, n_col + 1):
        cel = ws.cell(row=r, column=c)
        cel.font = F_BLD
        cel.border = BORDER
    ws.cell(row=r, column=1).value = label


def pct_num(p) -> float | None:
    """Converte % do JSON (ex: -33.87) em valor fracionario p/ formatar como %."""
    return None if p is None else p / 100.0


def aplicar_delta(cel, delta: float | None, bom_negativo: bool = True):
    """Pinta a celula de delta conforme semantica: verde quando (nao) foi bom."""
    if delta is None:
        cel.value = "—"
        return
    up = delta > 0
    bom = (not up) if bom_negativo else up
    cel.value = f"{'' if up else '−'}{numerico(abs(delta))}"
    cel.font = F_POS if bom else F_NEG


def main() -> None:
    d = extrair_dados()
    jul = d["consumoTend"][8]
    ago = d["consumoTend"][9]
    pj = d["por_mes"]["JULHO"]
    pa = d["por_mes"]["AGOSTO"]

    wb = Workbook()

    total_v = dict(zip(d["meses"], d["tendencia"]))  # pragma: no cover

    # =====================================================================
    # ABA 1 — RESUMO
    # =====================================================================
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:H1")
    ws["A1"] = "RESUMO — JULHO × AGOSTO (dados do Dashboard de Frota Carajás)"
    ws["A1"].font = F_TIT
    ws["A1"].fill = FILL_TIT
    ws["A1"].alignment = AL_C
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Métrica"
    ws["B3"] = "JULHO"
    ws["C3"] = "AGOSTO"
    ws["D3"] = "Δ (absoluto)"
    ws["E3"] = "Δ (%)"
    estilo_cabecalho(ws, 3, 5)

    met = [
        ("KM total", f_km(jul["km"]), f_km(ago["km"]), ago["km"] - jul["km"],
         (ago["km"] - jul["km"]) / jul["km"]),
        ("Veículos no mês", jul["veic"], ago["veic"], ago["veic"] - jul["veic"],
         (ago["veic"] - jul["veic"]) / jul["veic"]),
        ("Litros totais (L)", numerico(jul["litrosD"] + jul["litrosG"]),
         numerico(ago["litrosD"] + ago["litrosG"]),
         (ago["litrosD"] + ago["litrosG"]) - (jul["litrosD"] + jul["litrosG"]),
         ((ago["litrosD"] + ago["litrosG"]) - (jul["litrosD"] + jul["litrosG"]))
         / (jul["litrosD"] + jul["litrosG"])),
        ("CO₂ total (kgCO₂e)", numerico(jul["total"]), numerico(ago["total"]),
         ago["total"] - jul["total"], (ago["total"] - jul["total"]) / jul["total"]),
        ("Média por veículo (km)", f_km(round(jul["km"] / jul["veic"])),
         f_km(round(ago["km"] / ago["veic"])),
         ago["km"] / ago["veic"] - jul["km"] / jul["veic"],
         (ago["km"] / ago["veic"] - jul["km"] / jul["veic"]) / (jul["km"] / jul["veic"])),
    ]
    r = 4
    for nome, vj, va, delta, frac in met:
        ws.cell(row=r, column=1, value=nome).font = F_BLD
        ws.cell(row=r, column=2, value=vj)
        ws.cell(row=r, column=3, value=va)
        if delta is not None:
            aplicar_delta(ws.cell(row=r, column=4), delta)
            ws.cell(row=r, column=5).value = frac
            ws.cell(row=r, column=5).number_format = "+0.0%;-0.0%"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
            if c in (2, 3):
                ws.cell(row=r, column=c).alignment = AL_C
        r += 1

    ws["A9"] = "Frota operacional"
    ws["B9"] = "6 de 10 (60%)"
    ws["C9"] = "9 de 9 (100%)"
    ws["A10"] = "Em manutenção"
    ws["B10"] = "3"
    ws["C10"] = "0"
    for rr in (9, 10):
        ws.cell(row=rr, column=1).font = F_BLD
        for c in (1, 2, 3):
            ws.cell(row=rr, column=c).border = BORDER

    # grafico de barras agrupado
    dados_g = [
        ("KM (km)", jul["km"], ago["km"], "#F59E0B", "#2563EB"),
        ("Litros (L)", jul["litrosD"] + jul["litrosG"],
         ago["litrosD"] + ago["litrosG"], "#F59E0B", "#2563EB"),
        ("CO₂ (kg)", jul["total"], ago["total"], "#F59E0B", "#2563EB"),
    ]
    base = 12
    for i, (rot, vj, va, cj, ca) in enumerate(dados_g):
        rr = base + i * 3
        ws.cell(row=rr, column=2, value="JULHO")
        ws.cell(row=rr, column=3, value=vj)
        ws.cell(row=rr + 1, column=2, value="AGOSTO")
        ws.cell(row=rr + 1, column=3, value=va)
        grafico = BarChart()
        grafico.type = "col"
        grafico.style = 10
        grafico.width = 7.5
        grafico.height = 5
        grafico.title = rot
        grafico.y_axis.title = rot.split(" ")[0]
        grafico.add_data(
            Reference(ws, min_col=3, min_row=rr, max_row=rr + 1),
            titles_from_data=False)
        grafico.set_categories(Reference(ws, min_col=2, min_row=rr, max_row=rr + 1))
        if i == 0:
            grafico.dataLabels = DataLabelList()
            grafico.dataLabels.showVal = True
        ws.add_chart(grafico, f"G{base + i * 3}")

    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 18
    for col in "FGHI":
        ws.column_dimensions[col].width = 14

    # =====================================================================
    # ABA 2 — COMBUSTÍVEL
    # =====================================================================
    ws2 = wb.create_sheet("Combustível")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "COMBUSTÍVEL — Diesel e Gasolina (litros e CO₂)"
    ws2["A1"].font = F_TIT
    ws2["A1"].fill = FILL_TIT
    ws2["A1"].alignment = AL_C
    ws2.row_dimensions[1].height = 30

    ws2["A3"] = "Combustível"
    ws2["B3"] = "Litros JULHO"
    ws2["C3"] = "Litros AGOSTO"
    ws2["D3"] = "Δ litros"
    ws2["E3"] = "Δ litros (%)"
    ws2["F3"] = "CO₂ JULHO (kg)"
    ws2["G3"] = "CO₂ AGOSTO (kg)"
    ws2["H3"] = "Δ CO₂ (kg)"
    estilo_cabecalho(ws2, 3, 8)

    comb = [
        ("Diesel", jul["litrosD"], ago["litrosD"], jul["diesel"], ago["diesel"]),
        ("Gasolina", jul["litrosG"], ago["litrosG"], jul["gasolina"], ago["gasolina"]),
    ]
    r = 4
    for nome, lj, la, cj, ca in comb:
        ws2.cell(row=r, column=1, value=nome).font = F_BLD
        ws2.cell(row=r, column=2, value=lj)
        ws2.cell(row=r, column=3, value=la)
        aplicar_delta(ws2.cell(row=r, column=4), la - lj)
        ws2.cell(row=r, column=5).value = pct_num((la - lj) / lj * 100)
        ws2.cell(row=r, column=5).number_format = "+0.0%;-0.0%"
        ws2.cell(row=r, column=6, value=cj)
        ws2.cell(row=r, column=7, value=ca)
        aplicar_delta(ws2.cell(row=r, column=8), ca - cj)
        for c in range(1, 9):
            ws2.cell(row=r, column=c).border = BORDER
            if c > 1:
                ws2.cell(row=r, column=c).number_format = '#,##0.0'
                ws2.cell(row=r, column=c).alignment = AL_C
        ws2.row_dimensions[r].height = 20
        r += 1

    ws2.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws2.column_dimensions[col].width = 16

    # grafico CO2 por combustivel
    ws2["A10"] = "Atual"
    ws2["B10"] = "Mês"
    ws2["C10"] = "Diesel (kg)"
    ws2["D10"] = "Gasolina (kg)"
    ws2["A11"] = "Diesel"
    ws2["A12"] = "Gasolina"
    ws2["B11"] = "JULHO"
    ws2["B12"] = "AGOSTO"
    ws2["C11"] = jul["diesel"]
    ws2["C12"] = ago["diesel"]
    ws2["D11"] = jul["gasolina"]
    ws2["D12"] = ago["gasolina"]
    g2 = BarChart()
    g2.type = "col"
    g2.style = 10
    g2.width = 12
    g2.height = 7
    g2.title = "CO₂ (kg) por combustível — JULHO × AGOSTO"
    g2.add_data(Reference(ws2, min_col=3, min_row=11, max_col=4, max_row=12),
                titles_from_data=True, from_rows=False)
    g2.set_categories(Reference(ws2, min_col=2, min_row=11, max_row=12))
    g2.dataLabels = DataLabelList()
    g2.dataLabels.showVal = True
    ws2.add_chart(g2, "G10")

    # pizza participação Diesel x Gasolina em AGOSTO
    ws2["A21"] = "Participação no CO₂ de AGOSTO"
    ws2["A22"] = "Diesel"
    ws2["B22"] = ago["diesel"]
    ws2["A23"] = "Gasolina"
    ws2["B23"] = ago["gasolina"]
    pz = PieChart()
    pz.title = "Participação no CO₂ — AGOSTO"
    pz.width = 7.5
    pz.height = 7
    pz.add_data(Reference(ws2, min_col=2, min_row=22, max_row=23), titles_from_data=False)
    pz.set_categories(Reference(ws2, min_col=1, min_row=22, max_row=23))
    pz.dataLabels = DataLabelList()
    pz.dataLabels.showPercent = True
    for i, cor in enumerate(["#F59E0B", "#EF4444"]):
        pz.series[0].data_points.append(DataPoint(idx=i, spPr=GraphicalProperties(solidFill=cor)))
    ws2.add_chart(pz, "A26")

    # =====================================================================
    # ABA 3 — VEÍCULOS
    # =====================================================================
    ws3 = wb.create_sheet("Veículos")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "VEÍCULOS — KM e CO₂ de AGOSTO (comparado a JULHO)"
    ws3["A1"].font = F_TIT
    ws3["A1"].fill = FILL_TIT
    ws3["A1"].alignment = AL_C
    ws3.row_dimensions[1].height = 30

    ws3["A3"] = "Veículo"
    ws3["B3"] = "Modelo"
    ws3["C3"] = "Combustível"
    ws3["D3"] = "Km no mês"
    ws3["E3"] = "Litros"
    ws3["F3"] = "CO₂ AGOSTO (kg)"
    ws3["G3"] = "CO₂ JULHO (kg)"
    ws3["H3"] = "Δ CO₂ (%)"
    estilo_cabecalho(ws3, 3, 8)

    jul_por_tag = {r[0]: r[5] for r in d["consumo"]["JULHO"]}
    rows = sorted(d["consumo"]["AGOSTO"], key=lambda r: -r[5])
    r = 4
    for tag, modelo, comb, km, litros, co2 in rows:
        ws3.cell(row=r, column=1, value=tag).font = F_BLD
        ws3.cell(row=r, column=2, value=modelo)
        ws3.cell(row=r, column=3, value=comb)
        ws3.cell(row=r, column=4, value=km)
        ws3.cell(row=r, column=5, value=litros)
        ws3.cell(row=r, column=6, value=co2)
        ant = jul_por_tag.get(tag)
        if ant:
            ws3.cell(row=r, column=7, value=ant)
            pct = (co2 - ant) / ant * 100 / 100
            ws3.cell(row=r, column=8).value = pct
            ws3.cell(row=r, column=8).number_format = "+0.0%;-0.0%"
            cel = ws3.cell(row=r, column=8)
            cel.font = F_POS if pct <= 0 else F_NEG
        else:
            ws3.cell(row=r, column=7, value="—")
            ws3.cell(row=r, column=8, value="—")
        for c in range(1, 9):
            ws3.cell(row=r, column=c).border = BORDER
            if c in (4, 5, 6, 7, 8):
                ws3.cell(row=r, column=c).number_format = '#,##0.0'
                ws3.cell(row=r, column=c).alignment = AL_C
        r += 1

    for col, w in zip("ABCDEFGH", (10, 18, 13, 12, 10, 15, 14, 12)):
        ws3.column_dimensions[col].width = w

    ws3["A15"] = "Atual"
    ws3["B15"] = "Veículo"
    ws3["C15"] = "CO₂ AGO (kg)"
    for idx, (tag, _, _, _, _, co2) in enumerate(rows):
        rr = 16 + idx
        ws3.cell(row=rr, column=2, value=tag)
        ws3.cell(row=rr, column=3, value=co2)
    g3 = BarChart()
    g3.type = "col"
    g3.style = 10
    g3.width = 14
    g3.height = 8
    g3.title = "Top CO₂ por veículo — AGOSTO"
    g3.add_data(Reference(ws3, min_col=3, min_row=16, max_row=16 + len(rows)),
                titles_from_data=False, from_rows=True)
    g3.set_categories(Reference(ws3, min_col=2, min_row=16, max_row=16 + len(rows)))
    ws3.add_chart(g3, "E15")

    # =====================================================================
    # ABA 4 — MESES (TENDÊNCIA)
    # =====================================================================
    ws4 = wb.create_sheet("Meses")
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "KM TOTAL POR MÊS (toda a frota)"
    ws4["A1"].font = F_TIT
    ws4["A1"].fill = FILL_TIT
    ws4["A1"].alignment = AL_C
    ws4.row_dimensions[1].height = 30

    ws4["A3"] = "Mês"
    ws4["B3"] = "KM total"
    ws4["C3"] = "Veículos"
    ws4["D3"] = "CO₂ (kg)"
    ws4["E3"] = "CO₂ Diesel+Gasolina"
    estilo_cabecalho(ws4, 3, 5)

    co2_por_mes = {c["mes"]: c for c in d["consumoTend"]}
    r = 4
    for t in d["tendencia"]:
        ws4.cell(row=r, column=1, value=t["mes"]).font = F_BLD
        ws4.cell(row=r, column=2, value=t["km"])
        ws4.cell(row=r, column=3, value=t["veic"])
        c = co2_por_mes.get(t["mes"])
        if c:
            ws4.cell(row=r, column=4, value=c["total"])
            ws4.cell(row=r, column=5,
                     value=f'{numerico(c["diesel"])} + {numerico(c["gasolina"])}')
        else:
            ws4.cell(row=r, column=4, value="—")
            ws4.cell(row=r, column=5, value="—")
        for cc in range(1, 6):
            ws4.cell(row=r, column=cc).border = BORDER
        r += 1

    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 24

    ws4["B16"] = "Mês"
    ws4["C16"] = "KM"
    for idx, t in enumerate(d["tendencia"]):
        rr = 17 + idx
        ws4.cell(row=rr, column=2, value=t["mes"])
        ws4.cell(row=rr, column=3, value=t["km"])
    g4 = BarChart()
    g4.type = "col"
    g4.style = 10
    g4.width = 18
    g4.height = 8
    g4.title = "KM total por mês"
    g4.add_data(Reference(ws4, min_col=3, min_row=17, max_row=17 + len(d["tendencia"]) - 1),
                titles_from_data=False, from_rows=True)
    g4.set_categories(Reference(ws4, min_col=2, min_row=17, max_row=17 + len(d["tendencia"]) - 1))
    ws4.add_chart(g4, "E16")

    # =====================================================================
    # ABA 5 — TEXTOS
    # =====================================================================
    ws5 = wb.create_sheet("Textos da Apresentação")
    ws5.merge_cells("A1:B1")
    ws5["A1"] = "TEXTOS PRONTOS PARA APRESENTAÇÃO — JULHO × AGOSTO"
    ws5["A1"].font = F_TIT
    ws5["A1"].fill = FILL_TIT
    ws5["A1"].alignment = AL_C
    ws5.row_dimensions[1].height = 30

    textos = [
        ("FALA DE ABERTURA (1 min)",
         "\"Em AGOSTO a frota andou 12.727 km — contra 19.244 km em JULHO, queda de "
         "6.517 km (−33,9%) com 9 veículos ativos (JULHO teve 10). Essa redução derrubou "
         "o consumo de combustível e, direto, as emissões de CO₂. Menos rodagem, menos "
         "custo, menos emissão, menos exposição dos motoristas na estrada.\""),
        ("SLIDE 1 — Quadro comparativo",
         "KM total: 19.244 km → 12.727 km (Δ −6.517 km, −33,9%)  •  Veículos: 10 → 9 (−10,0%)  "
         "•  Litros totais: 1.991,45 L → 1.297,38 L (Δ −694,07 L, −34,9%)  •  CO₂ total: "
         "5.251,89 kg → 3.388,59 kg (Δ −1.863,30 kg, −35,5%)  •  Média por veículo: "
         "1.924,4 km → 1.414 km."),
        ("SLIDE 2 — Quilometragem (o que caiu)",
         "Média por veículo caiu de 1.924,4 km para 1.414 km (≈ −26%). TN 78 foi o que "
         "mais rodou em AGOSTO: 2.735 km (3.829 km em JULHO). TN 01 (Pulse) subiu: "
         "1.955 km contra 1.802 km. TN 76 saiu da operação (0 km em JULHO). Frota ficou "
         "100% operacional em AGOSTO (9 de 9), contra 60% em JULHO."),
        ("SLIDE 3 — Combustível (Diesel cai forte, Gasolina sobe)",
         "DIESEL: 1.827,63 L → 1.127,38 L (Δ −700,25 L, −38,3%) e CO₂ 4.898,04 kg → "
         "3.021,39 kg (Δ −1.876,65 kg, −38,3%).  GASOLINA: 163,82 L → 170,00 L (Δ +6,18 L, "
         "+3,8%) e CO₂ 353,85 kg → 367,20 kg (Δ +13,35 kg, +3,8%). Participação: Diesel "
         "de 93% para 89% das emissões; Gasolina de 7% para 11%. O que aumentou: apenas a "
         "gasolina do veículo leve (TN 01), que rodou +153 km no mês."),
        ("SLIDE 4 — Emissões de CO₂",
         "CO₂ evitado em AGOSTO: ~1,86 tonelada vs JULHO (−35,5%). Top emissores de AGOSTO: "
         "TN 78 com 771,6 kg (−28,6%), TN 29 401,2 kg (−55,9%), TN 04 382,3 kg (−14,9%), "
         "TN 72 374,9 kg (+3,5%), TN 01 367,2 kg (+3,8%). TN 78 sozinha ≈ 23% do CO₂ do mês "
         "e, ainda assim, emitiu ~309 kg a menos que em JULHO."),
        ("SLIDE 5 — Benefícios da redução",
         "1) Custo de combustível: −694 litros no mês. 2) Emissões: −1.863,30 kg CO₂e, apoia "
         "metas ESG. 3) Desgaste da frota: menos km = menos manutenção, pneus e revisões. "
         "4) Operação: frota 100% operacional em AGOSTO, sem veículos em manutenção nem "
         "mobilização."),
        ("SLIDE 6 — Exposição dos motoristas",
         "−6.517 km rodados significa centenas de horas a menos ao volante no mês: menos "
         "exposição a acidentes, estradas precárias e fadiga. Motoristas de caminhonete "
         "(frota Diesel) foram os mais beneficiados — concentraram 94%+ do km e tiveram as "
         "maiores reduções (ex.: TN 78 −1.094 km, TN 29 −1.802 km, TN 74 −1.582 km). Dias de "
         "operação mais leves → turnos mais curtos, mais segurança e qualidade de vida."),
        ("FALA DE ENCERRAMENTO",
         "\"AGOSTO entregou menos km (−33,9%), menos combustível (−34,9%) e menos CO₂ "
         "(−35,5%), mantendo a frota 100% operacional. O único aumento foi a gasolina do "
         "veículo leve (+3,8%), irrelevante no total. É o retrato de uma operação mais "
         "enxuta, mais segura e mais sustentável.\""),
    ]
    ws5["A3"] = "Seção"
    ws5["B3"] = "Texto"
    estilo_cabecalho(ws5, 3, 2)
    r = 4
    for titulo, corpo in textos:
        ws5.cell(row=r, column=1, value=titulo).font = F_BLD
        ws5.cell(row=r, column=2, value=corpo).alignment = AL_L
        ws5.row_dimensions[r].height = 66
        for c in (1, 2):
            ws5.cell(row=r, column=c).border = BORDER
        ws5.cell(row=r, column=1).alignment = AL_L
        r += 1

    ws5.column_dimensions["A"].width = 34
    ws5.column_dimensions["B"].width = 110

    wb.save(SAIDA)
    print(f"Gerado: {SAIDA} ({SAIDA.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()